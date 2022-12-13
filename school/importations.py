from datetime import datetime, date

from django.contrib.auth.hashers import make_password
from main.models import *
import threading
import openpyxl
from django.db import transaction, IntegrityError


def ReferenceMaker(sigle):
    time = datetime.now()
    datepart = str(date.today().year) + str(date.today().month) + str(date.today().day)
    reference = sigle + str(datepart) + "-" + str(time.strftime("%H%M%S"))
    return reference


#  gradurations imporation
class GraduationsLevelsImportation(threading.Thread):
    def __init__(self, importationId):
        self.importationId = importationId
        threading.Thread.__init__(self)

    def run(self):
        status = True
        report = ""
        importation = Importation.objects.get(id=self.importationId)
        document = openpyxl.load_workbook(importation.file.path)
        sheets = document.sheetnames
        # Verifier si il y a des feuilles dans le fichier
        if not sheets:
            status = False
            report = "Aucune donné à enregistrer"
        else:
            print("Debut de l'importation")
            importation.importationStatus = "En cours de traitement"
            importation.save()
            for sheet in sheets:
                try:
                    with transaction.atomic():
                        # Si le niveaux n existe pas alors le creer
                        if not GraduationLevel.objects.filter(wording__exact=sheet).exists():
                            GraduationLevel.objects.create(reference=ReferenceMaker(sheet + '-'), wording=sheet,
                                                           Importation_id=self.importationId)
                        # Recuperer le niveau concerne
                        graduationLevel = GraduationLevel.objects.get(wording__exact=sheet)
                        currentSheet = document[sheet]
                        for i in range(2, currentSheet.max_row + 1):
                            for j in range(1, currentSheet.max_column + 1):
                                # Si le degre d etude n existe pas alors le creer
                                if currentSheet.cell(i, 1).value:  # Verifier si le sigle existe deja pour ce degre
                                    if not SusGraduationLevel.objects.filter(sigle__exact=currentSheet.cell(i, 1).value,
                                                                             GraduationLevel=graduationLevel).exists():
                                        SusGraduationLevel.objects.create(GraduationLevel=graduationLevel,
                                                                          Importation_id=self.importationId,
                                                                          sigle=currentSheet.cell(i, 1).value,
                                                                          wording=currentSheet.cell(i, 2).value)
                        report = "Importation éffectuée"
                except:
                    report = "Erreur lors du traitement"

            importation = Importation.objects.get(id=self.importationId)
            importation.importationStatus = "Traitement terminé"
            importation.importationReport = report
            importation.save()
            print("Importation terminé")


# classes importations
class ClasseImportation(threading.Thread):
    def __init__(self, importationId):
        self.importationId = importationId
        threading.Thread.__init__(self)

    def run(self):
        status = True
        report = ""
        importation = Importation.objects.get(id=self.importationId)
        document = openpyxl.load_workbook(importation.file.path)
        sheets = document.sheetnames
        # Verifier si il y a des feuilles dans le fichier
        if not sheets:
            status = False
            report = "Aucune donné à enregistrer"
        else:
            print("Debut de l'importation")
            importation.importationStatus = "En cours de traitement"
            importation.save()
            for sheet in sheets:
                try:
                    with transaction.atomic():
                        if SusGraduationLevel.objects.filter(sigle__exact=sheet).exists():
                            # Recuperer le niveau concerne
                            susgraduationlevel = SusGraduationLevel.objects.get(sigle__exact=sheet)
                            currentSheet = document[sheet]
                            for i in range(2, currentSheet.max_row + 1):
                                for j in range(1, currentSheet.max_column + 1):
                                    if currentSheet.cell(i, 1).value:
                                        if not Classe.objects.filter(SusGraduationLevel=susgraduationlevel,
                                                                     School=importation.School,
                                                                     wording=currentSheet.cell(i, 1).value).exists():
                                            Classe.objects.create(SusGraduationLevel=susgraduationlevel,
                                                                  School=importation.School,
                                                                  Importation_id=self.importationId,
                                                                  reference=ReferenceMaker(
                                                                      currentSheet.cell(i, 1).value + '-'),
                                                                  wording=currentSheet.cell(i, 1).value,
                                                                  capacity=currentSheet.cell(i, 2).value)
                                report = "Importation éffectuée"

                except IntegrityError:
                    report = "Erreur lors du traitement"

            importation = Importation.objects.get(id=self.importationId)
            importation.importationStatus = "Traitement terminé"
            importation.importationReport = report
            importation.save()
            print("Importation terminé")


# Students importation
class StudentsImportation(threading.Thread):
    def __init__(self, importationId):
        self.importationId = importationId
        threading.Thread.__init__(self)

    def run(self):
        status = True
        report = ""
        importation = Importation.objects.get(id=self.importationId)
        document = openpyxl.load_workbook(importation.file.path)
        sheets = document.sheetnames

        # Enregistrement des inscription
        for sheet in sheets:
            print("Début de l'importation")
            importation.importationStatus = "En cours de traitement"
            importation.save()
            currentSheet = document[sheet]
            for i in range(2, currentSheet.max_row + 1):
                status = True
                # Verifier le matricule n est pas vide dans le fichier et si il n est deja pas attribue a un apprenant
                if currentSheet.cell(i, 1).value:
                    if User.objects.filter(matricule__exact=currentSheet.cell(i, 1).value).exists():
                        status = False
                if currentSheet.cell(i, 12).value:
                    if User.objects.filter(email__exact=currentSheet.cell(i, 12).value).exists():
                        status = False
                if currentSheet.cell(i, 13).value:
                    if User.objects.filter(phoneNumber__exact=currentSheet.cell(i, 13).value).exists():
                        status = False
                if status:
                    try:
                        with transaction.atomic():
                            matricule = currentSheet.cell(i, 1).value
                            lastName = currentSheet.cell(i, 2).value
                            firstName = currentSheet.cell(i, 3).value
                            gender = currentSheet.cell(i, 4).value
                            birthDate = currentSheet.cell(i, 5).value
                            birthCountry = currentSheet.cell(i, 6).value
                            birthTown = currentSheet.cell(i, 7).value
                            nationality = currentSheet.cell(i, 8).value
                            livingCountry = currentSheet.cell(i, 9).value
                            livingTown = currentSheet.cell(i, 10).value
                            address = currentSheet.cell(i, 11).value
                            email = currentSheet.cell(i, 12).value
                            phoneNumber = currentSheet.cell(i, 13).value
                            searchField = matricule + " " + lastName + " " + firstName + " " + str(birthDate) + " " + email + " " + str(phoneNumber)
                            newUser = User.objects.create(reference=ReferenceMaker(currentSheet.cell(i, 1).value + '-'),
                                                          matricule=matricule, lastName=lastName.upper(),
                                                          firstName=firstName.title(), gender=gender,
                                                          avatar='avatar.png',
                                                          birthDate=birthDate,
                                                          birthCountry=birthCountry.upper(),
                                                          birthTown=birthTown.title(),
                                                          nationality=nationality.upper(),
                                                          livingCountry=livingCountry.upper(),
                                                          livingTown=livingTown.title(), address=address,
                                                          email=email, phoneNumber=phoneNumber,
                                                          password=make_password('00000', 'sha512'),
                                                          searchField=searchField,
                                                          Importation_id=self.importationId)
                            # Lier l etudiant a l ecole
                            Learner.objects.create(User_id=newUser.id, School_id=importation.School_id, startDate=date.today(), isActive=True)
                            # Inscription
                            if Classe.objects.filter(School_id=importation.School_id, wording__exact=sheet).exists():
                                classe = Classe.objects.get(School_id=importation.School_id, wording__exact=sheet)
                                Inscription.objects.create(AcademicYear_id=importation.AcademicYear_id, inscriptionDate=date.today(),
                                                           SusGraduationLevel_id=classe.SusGraduationLevel_id,
                                                           Student_id=newUser.id, School_id=importation.School_id,
                                                           Classe_id=classe.id)
                            report = "Importation éffectuée"
                    except:
                        report = "Erreur lors du traitement"

            importation = Importation.objects.get(id=self.importationId)
            importation.importationStatus = "Traitement terminé"
            importation.importationReport = report
            importation.save()
            print("Importation terminée")
